# app/scenarios/routes.py
from flask import (render_template, redirect, url_for, flash, request,
                   current_app, jsonify, g, send_from_directory)
from flask_login import login_required, current_user 
from app import db
from app.scenarios import bp
from app.models import (Scenario, Consultant, ScenarioStep, ActionTypeEnum,
                        ReportTemplate, ColumnMapping, ProcessingRule, RuleOperatorEnum,
                        ExecutionLog, ExecutionStatusEnum)
from app.scenarios.forms import CreateScenarioForm, EditScenarioForm, ScenarioStepForm, ColumnMappingForm, ProcessingRuleForm
import os
import openpyxl
from sqlalchemy import func
import logging
from flask import request, jsonify
import uuid 
from werkzeug.utils import secure_filename 
from flask import current_app 
from .utils import get_excel_headers
from flask import request, jsonify, g 
from app.auth.decorators import token_required

# --- Routes ---


def get_template_headers(template_path):
    """Reads the first row from an Excel file to get headers."""
    headers = []
    if not template_path or not os.path.exists(template_path):
        logging.warning(f"Template path invalid or file not found: {template_path}")
        return headers # Return empty list if no path or file doesn't exist

    try:
        workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
        sheet = workbook.active # Get the first sheet
        # Iterate through the first row
        for cell in sheet[1]: # sheet[1] accesses the first row
            if cell.value: # Check if cell has a value
                headers.append(str(cell.value).strip()) # Add header, converting to string and stripping whitespace
        workbook.close() # Close workbook when done reading
        logging.debug(f"Read headers from template {template_path}: {headers}")
    except Exception as e:
        logging.error(f"Error reading headers from template {template_path}: {e}", exc_info=True)

@bp.route('/')
@bp.route('/list')
@login_required # Protect this route
def list_scenarios():
    page = request.args.get('page', 1, type=int)
    per_page = 15 
    search_term = request.args.get('search', '').strip()

    query = Scenario.query.filter_by(created_by_user_id=current_user.user_id)

    if search_term:
        # Simple search on name and URL
        query = query.filter(
            db.or_(
                Scenario.name.ilike(f'%{search_term}%'),
                Scenario.megara_url.ilike(f'%{search_term}%')
            )
        )
        logging.debug(f"Searching scenarios for term: '{search_term}'")

    pagination = query.order_by(Scenario.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    scenarios = pagination.items

    return render_template(
        'scenarios/list.html',
        title='My Monitoring Scenarios',
        scenarios=scenarios,
        pagination=pagination,
        search_term=search_term 
    )
@bp.route('/<int:scenario_id>')
@login_required
def view_scenario(scenario_id):
    """Displays details, steps, and mappings, populates mapping form."""
    scenario = Scenario.query.get_or_404(scenario_id)
    if scenario.created_by_user_id != current_user.user_id:
        flash('Permission denied.', 'danger'); return redirect(url_for('scenarios.list_scenarios'))

    mappings = scenario.column_mappings.all()
    mapping_form = ColumnMappingForm()

    # --- Populate Template Header Choices ---
    template_headers = []
    if scenario.template and scenario.template.file_path:
        template_headers = get_excel_headers(scenario.template.file_path)
        if template_headers:
             # Format choices for SelectField: (value, label)
             mapping_form.template_header.choices = [(h, h) for h in template_headers]
             # Add an empty default choice (optional but recommended)
             mapping_form.template_header.choices.insert(0, ('', '-- Select Template Header --'))
        else:
             flash("Could not read headers from the associated template file. Please check the file or upload a new one.", "warning")
             mapping_form.template_header.choices = [('', '-- No Headers Found --')]
    else:
        # If no template, disable the field or provide a single disabled choice
        mapping_form.template_header.choices = [('', '-- No Template Uploaded --')]
        # Optionally disable the submit button if no template? (Handled in template below)
    # --- End Populate Choices ---

    # --- Get Rules --- 
    rules = scenario.processing_rules.all() # Fetch existing rules

    # --- Instantiate Add Rule Form --- 
    rule_form = ProcessingRuleForm()

    return render_template(
        'scenarios/view.html',
        title=f'Scenario: {scenario.name}',
        scenario=scenario,
        mappings=mappings,
        mapping_form=mapping_form,
        has_template=bool(scenario.template),  
        rules=rules,          
        rule_form=rule_form 
    )
@bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_scenario():
    """Handles initial creation of a new scenario."""
    form = CreateScenarioForm()
    if form.validate_on_submit():
        # --- 1. Handle File Upload ---
        file = form.template_file.data # Get the uploaded file object
        template_record = None
        try:
            # Construct absolute upload path within instance folder
            upload_folder_rel = current_app.config.get('UPLOAD_FOLDER_RELATIVE', 'uploads/templates') # Get from config
            upload_folder_abs = os.path.join(current_app.instance_path, upload_folder_rel)
            os.makedirs(upload_folder_abs, exist_ok=True) # Create if not exists

            original_filename = secure_filename(file.filename)
            _, ext = os.path.splitext(original_filename)
            unique_filename = str(uuid.uuid4()) + ext # Create unique filename
            save_path = os.path.join(upload_folder_abs, unique_filename)

            file.save(save_path) # Save the physical file
            logging.info(f"Uploaded template for new scenario saved to: {save_path}")

            # Create ReportTemplate database record
            template_record = ReportTemplate(
                original_filename=original_filename,
                file_path=save_path # Store absolute path
            )
            db.session.add(template_record)
            db.session.flush() # To get the template_record.template_id

        except Exception as e:
            db.session.rollback()
            flash(f'Error uploading template file: {e}', 'danger')
            logging.error(f"Error uploading template for new scenario: {e}", exc_info=True)
            return render_template('scenarios/create_scenario_form.html', title='Create New Scenario', form=form)


        # --- 2. Create Scenario Record ---
        try:
            new_scenario = Scenario(
                name=form.name.data,
                megara_url=form.megara_url.data,
                schedule_cron=form.schedule_cron.data,
                created_by_user_id=current_user.user_id,
                report_template_id=template_record.template_id, # Link to the uploaded template
                custom_report_base_name=form.custom_report_base_name.data or None 
            )
            db.session.add(new_scenario)
            db.session.commit() # Commit both scenario and template
            flash(f'Scenario "{new_scenario.name}" and template uploaded successfully!', 'success')

            # --- 3. Redirect to Guidance Page ---
            return redirect(url_for('scenarios.start_recording_guidance',
                                    scenario_id=new_scenario.scenario_id))

        except Exception as e:
            db.session.rollback() # Rollback scenario and template creation
            flash(f'Error creating scenario: {e}', 'danger')
            logging.error(f"Error creating new scenario: {e}", exc_info=True)
            # Optionally delete the uploaded file if scenario creation fails here
            if save_path and os.path.exists(save_path):
                try: os.remove(save_path)
                except: logging.error(f"Could not remove orphaned template file: {save_path}")

    # --- Handle GET Request ---
    return render_template('scenarios/create_scenario_form.html', title='Create New Scenario', form=form)

# --- Recording Guidance Route ---
@bp.route('/<int:scenario_id>/record-guidance')
@login_required
def start_recording_guidance(scenario_id):
    scenario = db.session.get(Scenario, scenario_id) 
    if not scenario or scenario.created_by_user_id != current_user.user_id:
        flash("Scenario not found or permission denied.", "danger")
        return redirect(url_for('scenarios.list_scenarios'))

    return render_template(
        'scenarios/start_recording_guidance.html',
        title="Start Recording Steps",
        scenario=scenario
    )
@bp.route('/<int:scenario_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_scenario(scenario_id):
    scenario = Scenario.query.get_or_404(scenario_id)
    if scenario.created_by_user_id != current_user.user_id: # Check ownership
        flash('Permission denied.', 'danger'); return redirect(url_for('scenarios.list_scenarios'))

    # Pass scenario object to form only if needed for validation, not for pre-population here
    form = EditScenarioForm()

    if form.validate_on_submit(): # Process POST request
        # --- Handle File Upload ---
        file = form.template_file.data
        new_template_id = scenario.report_template_id # Keep existing template by default

        if file: # Check if a file was actually uploaded
            try:
                # Construct absolute upload path within instance folder
                upload_folder_rel = current_app.config['UPLOAD_FOLDER_RELATIVE']
                # app.instance_path is the absolute path to the instance folder
                upload_folder_abs = os.path.join(current_app.instance_path, upload_folder_rel)
                # Create directory if it doesn't exist
                os.makedirs(upload_folder_abs, exist_ok=True)

                # Sanitize and create unique filename
                original_filename = secure_filename(file.filename)
                # Use UUID to ensure uniqueness, keep original extension
                _, ext = os.path.splitext(original_filename)
                unique_filename = str(uuid.uuid4()) + ext
                save_path = os.path.join(upload_folder_abs, unique_filename)

                # Save the file
                file.save(save_path)
                logging.info(f"Uploaded template saved to: {save_path}")

                # --- Create new ReportTemplate record ---
                new_template = ReportTemplate(
                    original_filename=original_filename,
                    file_path=save_path # Store the absolute save path
                )
                db.session.add(new_template)
                db.session.flush() # Get the ID of the new template
                new_template_id = new_template.template_id # Use the new template ID
                flash(f'New template "{original_filename}" uploaded successfully.', 'info')

            except Exception as e:
                 db.session.rollback() # Rollback template creation if save fails
                 flash(f'Error uploading template file: {e}', 'danger')
                 logging.error(f"Error uploading template file: {e}", exc_info=True)
                 # Redirect back to edit form to show error
                 return render_template('scenarios/edit_form.html', title=f'Edit Scenario: {scenario.name}', form=form, scenario=scenario)
        # --- End File Upload Handling ---

        # --- Update Scenario fields ---
        try:
            scenario.name = form.name.data
            scenario.megara_url = form.megara_url.data
            scenario.schedule_cron = form.schedule_cron.data
            
            scenario.email_recipients = form.email_recipients.data
            scenario.upload_path = form.upload_path.data
            scenario.report_template_id = new_template_id # Assign new or existing template ID
            scenario.custom_report_base_name = form.custom_report_base_name.data or None

            db.session.commit() # Commit all scenario changes
            flash(f'Scenario "{scenario.name}" updated successfully!', 'success')
            return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating scenario details: {e}', 'danger')
            logging.error(f"Error updating scenario {scenario_id}: {e}", exc_info=True)


    elif request.method == 'GET': # Pre-populate form on GET request
        form.name.data = scenario.name
        form.megara_url.data = scenario.megara_url
        form.schedule_cron.data = scenario.schedule_cron
        
        form.email_recipients.data = scenario.email_recipients
        form.upload_path.data = scenario.upload_path
        form.custom_report_base_name.data = scenario.custom_report_base_name
        

    # Pass current template info for display
    current_template_filename = scenario.template.original_filename if scenario.template else None
    return render_template('scenarios/edit_form.html', title=f'Edit Scenario: {scenario.name}',
                           form=form, scenario=scenario, current_template=current_template_filename)
       

@bp.route('/<int:scenario_id>/delete', methods=['POST']) 
@login_required
def delete_scenario(scenario_id):
    """Handles deletion of a scenario."""
    scenario = Scenario.query.get_or_404(scenario_id)

    # Check ownership
    if scenario.created_by_user_id != current_user.user_id:
        flash('You do not have permission to delete this scenario.', 'danger')
        return redirect(url_for('scenarios.list_scenarios'))

    try:
        scenario_name = scenario.name # Get name before deleting
        # Deletion will cascade to steps, mappings, rules, logs due to cascade options in models
        db.session.delete(scenario)
        db.session.commit()
        flash(f'Scenario "{scenario_name}" and all associated data deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting scenario: {e}', 'danger')
        print(f"Error deleting scenario {scenario_id}: {e}") # Log error

    return redirect(url_for('scenarios.list_scenarios'))


# --- Add Processing Rule ---
@bp.route('/<int:scenario_id>/rules/add', methods=['POST'])
@login_required
def add_rule(scenario_id):
    """Handles adding a new processing rule via POST request."""
    scenario = Scenario.query.get_or_404(scenario_id)
    if scenario.created_by_user_id != current_user.user_id: # Check ownership
        flash('Permission denied.', 'danger'); return redirect(url_for('scenarios.list_scenarios'))

    form = ProcessingRuleForm() # Create form instance to validate POST data

    if form.validate_on_submit():
        try:
            # Validate operator requires value (unless IS_BLANK/IS_NOT_BLANK)
            operator_enum = RuleOperatorEnum[form.operator.data]
            if operator_enum not in [RuleOperatorEnum.IS_BLANK, RuleOperatorEnum.IS_NOT_BLANK] and not form.condition_value.data:
                 flash('Condition value is required for the selected operator.', 'warning')
                 # Redirect back to view which will show the flash message
                 return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

            new_rule = ProcessingRule(
                scenario_id=scenario.scenario_id,
                condition_column=form.condition_column.data,
                operator=operator_enum, # Store Enum member
                condition_value=form.condition_value.data,
                action_column=form.action_column.data,
                action_value=form.action_value.data
            )
            db.session.add(new_rule)
            db.session.commit()
            flash('Processing rule added successfully!', 'success')
        except KeyError:
             db.session.rollback()
             flash('Invalid operator selected.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding rule: {e}', 'danger')
            logging.error(f"Error adding rule for scenario {scenario_id}: {e}", exc_info=True)
    else:
        # Collect validation errors to show user
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in field '{getattr(form, field).label.text}': {error}", 'danger')

    # Redirect back to the scenario view page regardless of outcome
    return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))


# --- Delete Processing Rule ---
@bp.route('/<int:scenario_id>/rules/<int:rule_id>/delete', methods=['POST'])
@login_required
def delete_rule(scenario_id, rule_id):
    """Handles deleting an existing processing rule."""
    scenario = Scenario.query.get_or_404(scenario_id)
    rule = ProcessingRule.query.get_or_404(rule_id)

    # Check ownership and that rule belongs to the scenario
    if scenario.created_by_user_id != current_user.user_id or rule.scenario_id != scenario.scenario_id:
        flash('Permission denied or rule does not belong to this scenario.', 'danger')
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

    try:
        db.session.delete(rule)
        db.session.commit()
        flash('Processing rule deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting rule: {e}', 'danger')
        logging.error(f"Error deleting rule {rule_id}: {e}", exc_info=True)

    # Redirect back to the scenario view page
    return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))


# --- Add Column Mapping ---
@bp.route('/<int:scenario_id>/mappings/add', methods=['POST'])
@login_required
def add_mapping(scenario_id):
    """Handles adding a new column mapping via POST request."""
    scenario = Scenario.query.get_or_404(scenario_id)
    if scenario.created_by_user_id != current_user.user_id: # Check ownership
        flash('Permission denied.', 'danger'); return redirect(url_for('scenarios.list_scenarios'))

    form = ColumnMappingForm() # Create form instance to validate POST data

    if form.validate_on_submit():
        try:
            # Check if mapping for this scraped header already exists
            existing = ColumnMapping.query.filter_by(
                scenario_id=scenario.scenario_id,
                scraped_header=form.scraped_header.data
            ).first()
            if existing:
                flash(f"A mapping for scraped header '{form.scraped_header.data}' already exists.", 'warning')
            else:
                new_mapping = ColumnMapping(
                    scenario_id=scenario.scenario_id,
                    scraped_header=form.scraped_header.data,
                    template_header=form.template_header.data
                )
                db.session.add(new_mapping)
                db.session.commit()
                flash('Column mapping added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding mapping: {e}', 'danger')
            logging.error(f"Error adding mapping for scenario {scenario_id}: {e}", exc_info=True)
    else:
        # Collect validation errors to show user
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in field '{getattr(form, field).label.text}': {error}", 'danger')

    # Redirect back to the scenario view page regardless of outcome
    return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))


# --- Delete Column Mapping ---
@bp.route('/<int:scenario_id>/mappings/<int:mapping_id>/delete', methods=['POST'])
@login_required
def delete_mapping(scenario_id, mapping_id):
    """Handles deleting an existing column mapping."""
    scenario = Scenario.query.get_or_404(scenario_id)
    mapping = ColumnMapping.query.get_or_404(mapping_id)

    # Check ownership and that mapping belongs to the scenario
    if scenario.created_by_user_id != current_user.user_id or mapping.scenario_id != scenario.scenario_id:
        flash('Permission denied or mapping does not belong to this scenario.', 'danger')
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

    try:
        db.session.delete(mapping)
        db.session.commit()
        flash('Column mapping deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting mapping: {e}', 'danger')
        logging.error(f"Error deleting mapping {mapping_id}: {e}", exc_info=True)

    # Redirect back to the scenario view page
    return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))


# --- Add Scenario Step ---

@bp.route('/<int:scenario_id>/steps/add', methods=['GET', 'POST'])
@login_required
def add_step(scenario_id):
    """Handles adding a new step to a scenario."""
    scenario = Scenario.query.get_or_404(scenario_id)
    if scenario.created_by_user_id != current_user.user_id: # Check ownership
        flash('Permission denied.', 'danger'); return redirect(url_for('scenarios.list_scenarios'))

    form = ScenarioStepForm()

    # --- Handle POST Request ---
    if form.validate_on_submit():
        try:
            # Determine the next sequence order reliably
            last_order = db.session.query(func.max(ScenarioStep.sequence_order)).filter_by(scenario_id=scenario.scenario_id).scalar()
            next_order = (last_order or 0) + 1 # Start at 1 if no steps exist

            new_step = ScenarioStep(
                scenario_id=scenario.scenario_id,
                # Use the calculated next_order, ignore form input for sequence on create
                sequence_order=next_order,
                action_type=ActionTypeEnum[form.action_type.data], # Get Enum member
                selector=form.selector.data,
                value=form.value.data
            )
            db.session.add(new_step)
            db.session.commit()
            flash(f'New step (Order {next_order}) added successfully!', 'success')
            return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))
        except KeyError:
             # This happens if form.action_type.data is not a valid ActionTypeEnum key
             db.session.rollback()
             flash('Invalid action type selected.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding step: {e}', 'danger')
            logging.error(f"Error adding step for scenario {scenario_id}: {e}", exc_info=True)
            # Fall through to render form again with errors

    # --- Handle GET Request (or failed POST validation) ---
    # Calculate default sequence order for display on GET
    if request.method == 'GET':
        last_order = db.session.query(func.max(ScenarioStep.sequence_order)).filter_by(scenario_id=scenario.scenario_id).scalar()
        next_order = (last_order or 0) + 1
        form.sequence_order.data = next_order # Pre-populate for display

    # Render the form template
    return render_template('scenarios/step_form.html',
                           title=f'Add Step to Scenario: {scenario.name}',
                           form=form,
                           scenario=scenario,
                           # Make sequence readonly on add form
                           is_edit=False,
                           form_action=url_for('scenarios.add_step', scenario_id=scenario.scenario_id))

# --- Edit Scenario Step ---
@bp.route('/<int:scenario_id>/steps/<int:step_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_step(scenario_id, step_id):
    """Handles editing an existing scenario step."""
    scenario = Scenario.query.get_or_404(scenario_id)
    step = ScenarioStep.query.get_or_404(step_id)

    # Check ownership and that step belongs to the scenario
    if scenario.created_by_user_id != current_user.user_id or step.scenario_id != scenario.scenario_id:
        flash('Permission denied or step does not belong to this scenario.', 'danger')
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

    # Instantiate form, pre-populating with step data for GET requests
    form = ScenarioStepForm(obj=step)

    # --- Handle POST Request ---
    if form.validate_on_submit():
        try:
            # Check if sequence order is changing (potential future conflict)
            original_order = step.sequence_order
            new_order = form.sequence_order.data
            if original_order != new_order:
                logging.warning(f"Sequence order changed for step {step_id} from {original_order} to {new_order}. Re-ordering logic not implemented.")
                flash("Warning: Changing sequence order manually may lead to conflicts. Full re-ordering feature not yet implemented.", "warning")


            # Update the existing step object's attributes
            step.sequence_order = new_order
            step.action_type = ActionTypeEnum[form.action_type.data] # Get Enum member
            step.selector = form.selector.data
            step.value = form.value.data

            db.session.commit() # Commit the changes to the existing step
            flash('Step updated successfully!', 'success')
            return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))
        except KeyError:
            db.session.rollback()
            flash('Invalid action type selected.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating step: {e}', 'danger')
            logging.error(f"Error updating step {step_id}: {e}", exc_info=True)
            # Fall through to render form again with errors

    # Render the form template
    return render_template('scenarios/step_form.html',
                           title=f'Edit Step {step.sequence_order} in Scenario: {scenario.name}',
                           form=form,
                           scenario=scenario,
                           step=step, # Pass step for context
                           is_edit=True, # Flag for template logic
                           form_action=url_for('scenarios.edit_step', scenario_id=scenario.scenario_id, step_id=step.step_id))



# --- Delete Scenario Step ---
@bp.route('/<int:scenario_id>/steps/<int:step_id>/delete', methods=['POST']) # Only allow POST for deletion
@login_required
def delete_step(scenario_id, step_id):
    """Handles deleting an existing scenario step."""
    scenario = Scenario.query.get_or_404(scenario_id)
    step = ScenarioStep.query.get_or_404(step_id)

    # Check ownership and that step belongs to the scenario
    if scenario.created_by_user_id != current_user.user_id or step.scenario_id != scenario.scenario_id:
        flash('Permission denied or step does not belong to this scenario.', 'danger')
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

    try:
        step_order = step.sequence_order # Get order before deleting for message
        db.session.delete(step)
        db.session.commit()
        flash(f'Step {step_order} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting step: {e}', 'danger')
        print(f"Error deleting step {step_id}: {e}")

    # Redirect back to the scenario view page
    return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

# --- PROTECTED API ROUTE: Add Step ---
@bp.route('/api/scenarios/<int:scenario_id>/steps', methods=['POST'])
@token_required # Use the JWT token decorator
def api_add_step(scenario_id):
    """
    API endpoint to add a new step to a scenario, including optional
    mapping information (target header or target cell).
    Receives step data via JSON payload from the browser extension.
    """
    user = g.current_user # User object set by @token_required decorator
    if not user:
        return jsonify({"success": False, "message": "Authentication token invalid or user not found"}), 401

    scenario = db.session.get(Scenario, scenario_id)
    if not scenario:
        return jsonify({"success": False, "message": "Scenario not found"}), 404

    # Check ownership: ensure the scenario belongs to the authenticated user
    if scenario.created_by_user_id != user.user_id:
        return jsonify({"success": False, "message": "Permission denied to modify this scenario"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Invalid JSON payload"}), 400

    # Extract data from payload
    action_str = data.get('action_type')
    selector = data.get('selector') # For desktop, this will be JSON string of criteria
    value = data.get('value')       # Can be None for certain actions

    
    mapping_target_cell = data.get('mapping_target_cell')
    


    # Validate required fields
    if not action_str:
         return jsonify({"success": False, "message": "Missing 'action_type'"}), 400
    # Selector is not strictly required for all actions (e.g., WAIT_FOR_TIMEOUT, or NAVIGATE to a fixed URL)
    # but the form usually enforces it. 
    if not selector and action_str not in ['WAIT_FOR_TIMEOUT', 'NAVIGATE']: # Adjust as needed
         logging.warning(f"API: Adding step with action '{action_str}' but no selector provided.")
         # For NAVIGATE, if selector is empty, the runner uses scenario.megara_url or step.value if provided
         # For WAIT_FOR_TIMEOUT, selector is irrelevant.

    try:
        action_enum = ActionTypeEnum[action_str.upper()] # Convert string to Enum member
    except KeyError:
         return jsonify({"success": False, "message": f"Invalid action_type: '{action_str}'. Valid types are: {[e.name for e in ActionTypeEnum]}"}), 400

    try:
        # Determine the next sequence order for the new step
        last_order = db.session.query(func.max(ScenarioStep.sequence_order))\
            .filter_by(scenario_id=scenario.scenario_id)\
            .scalar()
        next_order = (last_order or 0) + 1

        new_step = ScenarioStep(
            scenario_id=scenario.scenario_id,
            sequence_order=next_order,
            action_type=action_enum,
            selector=selector if selector is not None else "", # Ensure selector is string, even if empty
            value=value,
            mapping_target_cell=mapping_target_cell
        )
        db.session.add(new_step)
        db.session.commit()


        # Return success and the details of the created step
        return jsonify({
            "success": True,
            "message": "Step added successfully",
            "step": {
                "step_id": new_step.step_id,
                "scenario_id": new_step.scenario_id,
                "sequence_order": new_step.sequence_order,
                "action_type": new_step.action_type.name, # Send enum name
                "selector": new_step.selector,
                "value": new_step.value,
                "mapping_target_cell": new_step.mapping_target_cell,
                "created_at": new_step.created_at.isoformat() if new_step.created_at else None
            }
        }), 201 # HTTP 201 Created status

    except Exception as e:
        db.session.rollback()
        logging.error(f"API Error adding step for scenario {scenario_id}: {e}", exc_info=True)
        return jsonify({"success": False, "message": f"Internal server error while saving step: {str(e)}"}), 500

# --- PROTECTED API ROUTE: List Scenarios ---
@bp.route('/api/scenarios/list', methods=['GET'])
@token_required # Use the JWT token decorator
def api_list_scenarios():
    """API endpoint to list scenarios owned by the JWT-authenticated user."""
    # Access user from g instead of current_user
    user = g.current_user
    if not user: 
        return jsonify({"success": False, "message": "Authentication failed"}), 401
    try:
        scenarios = Scenario.query.filter_by(created_by_user_id=user.user_id).order_by(Scenario.name).all()
        output = [{"id": scenario.scenario_id, "name": scenario.name} for scenario in scenarios]
        return jsonify({"success": True, "scenarios": output})
    except Exception as e:
        logging.error(f"API Error listing scenarios for user {user.user_id}: {e}", exc_info=True)
        return jsonify({"success": False, "message": f"Internal server error: {e}"}), 500

@bp.route('/api/scenarios/<int:scenario_id>/template/preview', methods=['GET'])
@token_required # Ensure the user is authenticated and g.current_user is set
def api_get_template_preview(scenario_id):
    """
    API endpoint to get a preview of the scenario's report template.
    Allows specifying a sheet_name via query parameter.
    Defaults to the first active sheet if no sheet_name is provided.
    Generates a grid of empty cells up to max_rows/max_cols if actual data is less.
    Accepts optional query parameters: sheet_name, max_rows, max_cols.
    """
    user = g.current_user # User object from the @token_required decorator
    scenario = db.session.get(Scenario, scenario_id)

    # --- Validations ---
    if not scenario:
        return jsonify({"success": False, "message": "Scenario not found"}), 404
    if scenario.created_by_user_id != user.user_id:
        return jsonify({"success": False, "message": "Permission denied to access this scenario's template"}), 403
    if not scenario.template or not scenario.template.file_path:
        return jsonify({"success": False, "message": "No report template associated with this scenario or path is missing"}), 404
    if not os.path.exists(scenario.template.file_path):
        logging.error(f"API: Template file not found at path: {scenario.template.file_path} for scenario {scenario_id}")
        return jsonify({"success": False, "message": "Template file not found on server. Please re-upload."}), 500
    # --- End Validations ---

    workbook = None # Initialize for finally block
    try:
        # Load the workbook in read-only and data_only mode
        workbook = openpyxl.load_workbook(scenario.template.file_path, read_only=True, data_only=True)
        all_sheet_names = workbook.sheetnames # Get all available sheet names

        if not all_sheet_names:
             workbook.close() # Close before returning
             return jsonify({"success": False, "message": "The Excel template contains no sheets."}), 400

        # --- Sheet Selection Logic ---
        requested_sheet_name = request.args.get('sheet_name')
        sheet_to_use = None
        sheet_name_to_report = None

        if requested_sheet_name:
            if requested_sheet_name in all_sheet_names:
                sheet_to_use = workbook[requested_sheet_name]
                sheet_name_to_report = requested_sheet_name
                logging.info(f"API: Using requested sheet: '{requested_sheet_name}' for preview.")
            else:
                logging.warning(f"API: Requested sheet '{requested_sheet_name}' not found in template. Available: {all_sheet_names}. Defaulting to first active sheet.")
                sheet_to_use = workbook.active
                sheet_name_to_report = sheet_to_use.title
        else:
            # Default to the first active sheet if no sheet_name is provided
            sheet_to_use = workbook.active
            sheet_name_to_report = sheet_to_use.title
            logging.info(f"API: No sheet_name requested, using active sheet: '{sheet_name_to_report}' for preview.")
        # --- End Sheet Selection ---

        # --- Define preview range (how many rows/cols to construct the grid for) ---
        try:
            max_rows_to_show_in_preview = int(request.args.get('max_rows', 20)) # How many rows to construct in preview
            max_cols_to_show_in_preview = int(request.args.get('max_cols', 10)) # How many cols to construct in preview
        except ValueError:
            workbook.close() # Close before returning
            return jsonify({"success": False, "message": "Invalid max_rows or max_cols parameter."}), 400

        logging.info(f"API: Generating preview grid - Max Display Rows: {max_rows_to_show_in_preview}, Max Display Cols: {max_cols_to_show_in_preview}")

        preview_data = [] # List of lists for cell values

        for r_idx in range(1, max_rows_to_show_in_preview + 1): # Iterate up to max_rows_to_show
            current_row_data = []
            for c_idx in range(1, max_cols_to_show_in_preview + 1): # Iterate up to max_cols_to_show
                # Get cell directly by its coordinates from the selected sheet
                # openpyxl is 1-indexed for rows and columns
                cell = sheet_to_use.cell(row=r_idx, column=c_idx)
                # Store cell value as string, or empty string if cell is None or has no value
                current_row_data.append(str(cell.value) if cell.value is not None else "")
            preview_data.append(current_row_data)

        # Get actual sheet dimensions from openpyxl for context
        actual_sheet_max_row = sheet_to_use.max_row
        actual_sheet_max_col = sheet_to_use.max_column
        # Handle openpyxl reporting 1x1 for a truly blank sheet (just created, never touched)
        if actual_sheet_max_row == 1 and actual_sheet_max_col == 1 and sheet_to_use['A1'].value is None:
            actual_sheet_max_row = 0
            actual_sheet_max_col = 0

        workbook.close() # Close workbook after reading all necessary data

        message = "Preview data fetched successfully."
        # Check if all cells in the generated preview_data are empty strings
        if not any(any(cell_val for cell_val in row) for row in preview_data):
            message = "Preview range appears to be empty or contains no values."
            if actual_sheet_max_row == 0 : # If the sheet itself is reported as empty
                message = "Template sheet appears to be completely empty."

        return jsonify({
            "success": True,
            "sheet_name": sheet_name_to_report,    # Name of the sheet used
            "all_sheet_names": all_sheet_names,   # List of all available sheets
            "preview_data": preview_data,         # The grid of cell values (strings)
            "fetched_rows": len(preview_data),    # Number of rows in preview_data (max_rows_to_show_in_preview)
            "fetched_cols": len(preview_data[0]) if preview_data and preview_data[0] else 0, # Number of cols in preview_data (max_cols_to_show_in_preview)
            "actual_rows_in_sheet": actual_sheet_max_row, # Max row from openpyxl for the sheet
            "actual_cols_in_sheet": actual_sheet_max_col, # Max col from openpyxl for the sheet
            "message": message
        })

    except Exception as e:
        logging.error(f"API Error getting template preview for scenario {scenario_id}: {e}", exc_info=True)
        # Ensure workbook is closed if it was opened and an error occurred
        if 'workbook' in locals() and workbook and not workbook.read_only: # read_only workbooks don't need explicit close on error as much
            try:
                workbook.close()
            except Exception as e_close:
                logging.error(f"Error closing workbook during exception handling: {e_close}")
        return jsonify({"success": False, "message": f"Error reading template preview: {str(e)}"}), 500

@bp.route('/logs/<int:log_id>/download_report') # Check this URL pattern
@login_required
def download_report(log_id): # Check this function name
    """Serves the report file associated with an execution log for download."""
    log_entry = db.session.get(ExecutionLog, log_id)

    if not log_entry:
        flash("Execution log not found.", "danger")
        return redirect(request.referrer or url_for('scenarios.list_scenarios'))

    scenario = log_entry.scenario
    if not scenario or scenario.created_by_user_id != current_user.user_id:
        flash("Permission denied to download this report.", "danger")
        return redirect(url_for('scenarios.list_scenarios'))

    report_path_str = log_entry.report_file_path
    if not report_path_str:
        flash("No report file associated with this execution log.", "warning")
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))

    # ... (Security checks for path traversal - keep these) ...
    # ... (os.path.exists and os.path.isfile checks - keep these) ...
    allowed_report_base_dir_instance = os.path.join(current_app.instance_path, current_app.config.get('REPORT_OUTPUT_DIR', 'generated_reports').lstrip('./').lstrip('.\\'))
    allowed_report_base_dir_project = os.path.abspath(current_app.config.get('REPORT_OUTPUT_DIR', './generated_reports'))
    report_abs_path = os.path.abspath(report_path_str)
    is_in_instance_path = report_abs_path.startswith(os.path.abspath(current_app.instance_path))
    is_in_project_report_path = report_abs_path.startswith(allowed_report_base_dir_project)
    if not (is_in_instance_path or is_in_project_report_path):
         if not (report_abs_path.startswith(allowed_report_base_dir_instance) or report_abs_path.startswith(allowed_report_base_dir_project)):
            flash("Invalid report file path.", "danger")
            logging.warning(f"Potential path traversal attempt or invalid report path for log {log_id}: {report_path_str}")
            return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))
    if not os.path.exists(report_abs_path) or not os.path.isfile(report_abs_path):
        flash(f"Report file not found on server at expected location: {os.path.basename(report_path_str)}", "danger")
        logging.error(f"Report file for log {log_id} not found at: {report_abs_path}")
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))


    try:
        directory_path = os.path.dirname(report_abs_path)
        file_name = os.path.basename(report_abs_path)
        logging.info(f"Attempting to serve file: Directory='{directory_path}', Filename='{file_name}'")
        return send_from_directory(
            directory_path,
            file_name,
            as_attachment=True
        )
    except Exception as e:
        flash("Error serving the report file.", "danger")
        logging.error(f"Error serving report for log {log_id}: {e}", exc_info=True)
        return redirect(url_for('scenarios.view_scenario', scenario_id=scenario.scenario_id))