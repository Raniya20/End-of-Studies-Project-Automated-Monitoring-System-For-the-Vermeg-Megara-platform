# app/runner.py

import os
import time
import logging
from datetime import datetime, time as dt_time, date as dt_date, timedelta # Ensure time, date are aliased if datetime.time/date used
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError, Page, expect
import pandas as pd
import numpy as np # Import numpy for NaN
import re # Import re for parsing remarks
import joblib # Import joblib for loading models
from sklearn.preprocessing import StandardScaler # Example preprocessor type hint
# from sklearn.ensemble import IsolationForest # Example model type hint
from sqlalchemy.orm import scoped_session, sessionmaker
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path # For easier path manipulation

from flask_mail import Message # Import Message class
from flask import current_app # To access app config like MAIL_DEFAULT_SENDER

# Import necessary components from your app
from app import db # Assumes running within Flask context or DB session configured
from app.models import (Scenario, ScenarioStep, ExecutionLog, MonitoringResult,
                        ProcessingRule, ActionTypeEnum, ExecutionStatusEnum, RuleOperatorEnum)

from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.impute import SimpleImputer
from ml_model.predictor import predict_anomalies, EXPECTED_FEATURE_NAMES, pipeline as ml_pipeline_from_predictor
# from sklearn.ensemble import IsolationForest # Type hint if needed

# --- Define Constants needed for Feature Engineering in this Runner ---
# (These should mirror what your training script uses to produce the columns in EXPECTED_FEATURE_NAMES,
#  or be the names of columns already present in the scraped DataFrame df_input)

# Standardized Column Names (from scraped/pre-processed data available in df_input)
STD_COL_PROCESS = 'Process'
STD_COL_PERIODICITY = 'Periodicity'
STD_COL_STATUS_TODAY = 'Status_Today' # Standardized status after rules
STD_COL_REMARKS_RAW = 'Remarks_Raw' # If you save raw remarks
STD_COL_TIME_EXPECTED_STR = 'Time_Expected_Str'
STD_COL_TIME_ACTUAL_STR = 'Time_Actual_Str'
STD_COL_REPORT_DATE = 'Report_Date' # Should be a datetime.date object in df_input
STD_COL_STATUS_YESTERDAY_RAW = 'Status_Yesterday' # Raw D-1 status from input

# Fixed monitoring time for PENDING duration calculation (example)
MONITORING_HOUR = 19
MONITORING_MINUTE = 0


# --- Helper Functions (used for feature engineering within the runner) ---
def parse_time_robust(time_input):
    if pd.isna(time_input): return None
    if isinstance(time_input, dt_time): return time_input
    if isinstance(time_input, datetime): return time_input.time()
    time_str = str(time_input).strip()
    formats_to_try = ["%H:%M", "%H:%M:%S", "%I:%M:%S %p", "%I:%M %p", "%H.%M.%S"]
    for fmt in formats_to_try:
        try: return datetime.strptime(time_str, fmt).time()
        except ValueError: continue
    logging.debug(f"Runner: Could not parse time: '{time_input}'")
    return None

def parse_count_from_remarks(remark): # If 'Extracted_Count' is an EXPECTED_FEATURE_NAME
    if not isinstance(remark, str): return np.nan
    match = re.search(r'\d{1,3}(?:[,\s]\d{3})*|\d+', remark)
    if match:
        try: return int(match.group(0).replace(',', '').replace(' ', ''))
        except: return np.nan
    return np.nan

def calculate_delay(report_date, expected_time_obj, actual_time_obj):
    if pd.isna(report_date) or expected_time_obj is None or actual_time_obj is None: return np.nan
    try:
        if isinstance(report_date, datetime): report_date = report_date.date()
        elif not isinstance(report_date, dt_date): report_date = pd.to_datetime(report_date).date()
        dt_expected = datetime.combine(report_date, expected_time_obj)
        dt_actual = datetime.combine(report_date, actual_time_obj)
        if dt_actual < dt_expected: dt_actual += timedelta(days=1)
        return (dt_actual - dt_expected).total_seconds() / 60.0
    except: return np.nan

# --- Configure basic logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - RUNNER - %(levelname)s - %(message)s')


class AutomationRunner:
    """
    Executes a monitoring scenario using Playwright based on configuration
    stored in the database. Includes data processing, anomaly detection structure,
    and report generation.
    """
    
    DEFAULT_TIMEOUT = 15000

    def __init__(self, scenario_id: int, app, headless: bool = True):
        self.scenario_id = scenario_id
        self.headless = headless
        self.app = app # Flask app instance for config and context
        self.scenario: Scenario | None = None
        self.log_entry: ExecutionLog | None = None
        self.playwright = None
        self.browser = None
        self.page: Page | None = None
        self.results = {} # Stores {label: DataFrame or str} from extraction
        self.processing_rules: list[ProcessingRule] = []
        # ML artifacts are not loaded here; predictor.py handles that.
        self.typical_durations = {}
        self.typical_durations_path = os.environ.get('TYPICAL_DURATIONS_PATH', 'typical_durations.json')


    def _setup(self):
        """Loads scenario, rules, creates log entry, initializes Playwright."""
        logging.info(f"Setting up runner for Scenario ID: {self.scenario_id}")
        self.scenario = db.session.get(Scenario, self.scenario_id)
        if not self.scenario: raise ValueError(f"Scenario ID {self.scenario_id} not found.")
        logging.info(f"Loaded Scenario: {self.scenario.name}")

        self.processing_rules = self.scenario.processing_rules.all()
        logging.info(f"Loaded {len(self.processing_rules)} processing rules.")

        self.log_entry = ExecutionLog(
            scenario_id=self.scenario_id,
            start_time=datetime.utcnow(), # Use timezone.utc if using tz-aware datetimes
            status=ExecutionStatusEnum.RUNNING
        )
        db.session.add(self.log_entry)
        db.session.flush()
        logging.info(f"Created Execution Log Entry ID: {self.log_entry.log_id}")

        logging.info(f"Launching Playwright (Headless: {self.headless})...")
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=self.headless)
        self.page = self.browser.new_page()
        self.page.set_default_timeout(self.DEFAULT_TIMEOUT)
        logging.info("Playwright browser page initialized.")
        # No _load_ml_model() call needed here.
        self._load_typical_durations()


    def _teardown(self, status: ExecutionStatusEnum, message: str | None = None):
        """Closes Playwright, updates log entry."""
        logging.info("Tearing down runner...")
        # 1. Close Playwright
        if self.browser and self.browser.is_connected():
            try: self.browser.close()
            except Exception as e: logging.error(f"Error closing browser: {e}")
            finally: logging.info("Playwright browser closed.") # Log even if error during close
        if self.playwright:
            try: self.playwright.stop()
            except Exception as e: logging.error(f"Error stopping playwright: {e}")
            finally: logging.info("Playwright context stopped.")

        # 2. Update ExecutionLog entry
        if self.log_entry:
            self.log_entry.end_time = datetime.utcnow()
            self.log_entry.status = status
            if message: self.log_entry.log_message = (self.log_entry.log_message + "\n---\n" + message if self.log_entry.log_message else message)
            try:
                db.session.commit() # Commit log changes
                logging.info(f"Execution Log ID {self.log_entry.log_id} updated. Final Status: {status.name}")
            except Exception as e:
                 logging.error(f"Failed to commit final Execution Log update for ID {self.log_entry.log_id}: {e}")
                 db.session.rollback()


    def _get_dynamic_value(self, value: str | None) -> str | None:
        """Replaces date markers like __TODAY__ or __YESTERDAY__."""
        if value == "__TODAY__": return datetime.now().strftime('%Y-%m-%d')
        elif value == "__YESTERDAY__": return (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        return value


    def _execute_step(self, step: ScenarioStep):
        """Executes a single step from the scenario."""
        action = step.action_type
        selector = step.selector
        raw_value = step.value
        value = self._get_dynamic_value(raw_value)
        logging.info(f"Executing Step {step.sequence_order}: {action.name} - Selector: '{selector}' RawValue: '{raw_value}' -> Value: '{value}'")
        if not self.page: raise RuntimeError("Playwright page is not available.")

        try:
            if action == ActionTypeEnum.NAVIGATE:
                target_url = value or self.scenario.megara_url
                if not target_url: raise ValueError("Navigation target URL is missing.")
                self.page.goto(target_url, wait_until='networkidle')
                logging.info(f"  Navigated to {target_url}")
            elif action == ActionTypeEnum.CLICK:
                locator = self.page.locator(selector)
                locator.wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
                locator.click()
                logging.info(f"  Clicked '{selector}'")
            elif action == ActionTypeEnum.TYPE:
                locator = self.page.locator(selector)
                locator.wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
                locator.fill(value if value is not None else "")
                logging.info(f"  Typed into '{selector}'")
            elif action == ActionTypeEnum.SELECT:
                if value is None: raise ValueError(f"Missing value for SELECT on step {step.sequence_order}")
                locator = self.page.locator(selector)
                locator.wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
                locator.select_option(value=value)
                logging.info(f"  Selected option with value '{value}' in '{selector}'")
            elif action == ActionTypeEnum.EXTRACT_ELEMENT:
                locator = self.page.locator(selector)
                locator.wait_for(state="attached", timeout=self.DEFAULT_TIMEOUT)
                text_content = locator.text_content()
                label = value if value else f"extracted_{step.step_id}"
                self.results[label] = text_content
                logging.info(f"  Extracted text from '{selector}' (Label: {label}): '{text_content[:100]}...'")
            elif action == ActionTypeEnum.EXTRACT_TABLE:
                label = value if value else f"table_{step.step_id}"
                logging.info(f"  Attempting table extraction for '{selector}' (Label: {label})...")
                table_locator = self.page.locator(selector)
                table_locator.wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
                table_html = table_locator.evaluate("element => element.outerHTML")
                try:
                    dfs = pd.read_html(table_html, flavor='lxml')
                    if dfs:
                        extracted_df = dfs[0]
                        self.results[label] = extracted_df
                        logging.info(f"  Successfully extracted table '{label}' with {len(extracted_df)} rows and {len(extracted_df.columns)} columns.")
                    else:
                        logging.warning(f"  Pandas read_html found no tables for selector '{selector}'.")
                        self.results[label] = pd.DataFrame()
                except ImportError:
                    logging.error("Pandas/lxml not installed for table extraction.")
                    self.results[label] = pd.DataFrame() # Store empty
                except ValueError as ve: # Handle cases where read_html finds no table
                    logging.warning(f"  Pandas read_html failed for selector '{selector}': {ve}")
                    self.results[label] = pd.DataFrame()
            elif action == ActionTypeEnum.WAIT_FOR_SELECTOR:
                target_state, wait_timeout = 'visible', self.DEFAULT_TIMEOUT
                if value:
                    parts = value.split(':', 1)
                    target_state = parts[0].lower()
                    if len(parts) > 1:
                        try: wait_timeout = int(parts[1])
                        except ValueError: logging.warning(f"Invalid timeout value '{parts[1]}'. Using default.")
                valid_states = ['attached', 'detached', 'visible', 'hidden']
                if target_state not in valid_states: raise ValueError(f"Invalid state '{target_state}' for WAIT.")
                logging.info(f"  Waiting for selector '{selector}' to be '{target_state}' (Timeout: {wait_timeout}ms)...")
                self.page.locator(selector).wait_for(state=target_state, timeout=wait_timeout)
                logging.info(f"  Selector '{selector}' is now '{target_state}'.")
            elif action == ActionTypeEnum.WAIT_FOR_TIMEOUT:
                try: timeout_ms = int(value) if value else 1000
                except (ValueError, TypeError): raise ValueError(f"Invalid timeout value '{value}' for WAIT_TIMEOUT.")
                logging.info(f"  Waiting for fixed timeout of {timeout_ms} ms...")
                self.page.wait_for_timeout(timeout_ms)
                logging.info("  Wait complete.")
            else:
                logging.warning(f"  Action type '{action.name}' not implemented yet.")
        except PlaywrightTimeoutError as e:
            logging.error(f"Timeout error on step {step.sequence_order} ({action.name}): {e}")
            raise
        except Exception as e:
            logging.error(f"Error executing step {step.sequence_order} ({action.name}): {e}", exc_info=True)
            raise


    def _perform_login(self):
        """Performs login to the target application using credentials from env vars if configured."""
        login_url = os.environ.get('MEGARA_LOGIN_URL')
        username = os.environ.get('MEGARA_USERNAME')
        password = os.environ.get('MEGARA_PASSWORD')
        if not login_url or not username or not password:
            logging.info("Login credentials/URL not fully configured in environment variables. Skipping login.")
            return
        if not self.page: raise RuntimeError("Login attempted before page initialization.")

        logging.info(f"Attempting login to {login_url}...")
        try:
            self.page.goto(login_url, wait_until='networkidle')
            user_selector = os.environ.get('MEGARA_USER_SELECTOR', "#username")
            pass_selector = os.environ.get('MEGARA_PASS_SELECTOR', "#password")
            submit_selector = os.environ.get('MEGARA_SUBMIT_SELECTOR', "button[type='submit']")
            post_login_wait_selector = os.environ.get('MEGARA_POST_LOGIN_SELECTOR', None)

            self.page.locator(user_selector).wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
            self.page.locator(user_selector).fill(username)
            logging.info("  Entered username.")
            self.page.locator(pass_selector).wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
            self.page.locator(pass_selector).fill(password)
            logging.info("  Entered password.")
            self.page.locator(submit_selector).wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
            self.page.locator(submit_selector).click()
            logging.info("  Clicked login button.")

            if post_login_wait_selector:
                 logging.info(f"  Waiting for post-login element: {post_login_wait_selector}")
                 self.page.locator(post_login_wait_selector).wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT * 2)
            else:
                 logging.info("  Waiting for network idle after login click...")
                 self.page.wait_for_load_state('networkidle', timeout=self.DEFAULT_TIMEOUT * 2)
            logging.info("Login completed (based on wait condition). Further verification recommended.")
        except PlaywrightTimeoutError as e:
            logging.error(f"Timeout during login process: {e}")
            raise
        except Exception as e:
            logging.error(f"Error during login process: {e}")
            raise


    def _apply_anomaly_detection(self, df_input: pd.DataFrame, label: str) -> pd.DataFrame:
        """
        Prepares features from df_input (scraped data after rules)
        and calls the ML predictor.
        """
        df_output = df_input.copy()
        if 'Is_Anomaly' not in df_output.columns: df_output['Is_Anomaly'] = False
        if 'Anomaly_Probability_Normal' not in df_output.columns: df_output['Anomaly_Probability_Normal'] = pd.NA

        # --- Use the imported alias 'ml_pipeline_from_predictor' ---
        if not self.scenario or not self.scenario.enable_anomalies: # If you keep this flag
            logging.debug(f"AD Skip for '{label}': Anomaly detection not enabled for this scenario.")
            return df_output
        if not ml_pipeline_from_predictor or not EXPECTED_FEATURE_NAMES: # Use correct variable name
            logging.warning(f"AD Skip for '{label}': ML pipeline or expected features not loaded by predictor module.")
            return df_output
        if df_output.empty:
            logging.info(f"AD Skip for '{label}': Input DataFrame is empty.")
            return df_output
        # --- END CORRECTION ---

        logging.info(f"Preparing data from '{label}' for anomaly detection (Model expects: {EXPECTED_FEATURE_NAMES})...")

        data_for_prediction_list = []
        for index, row in df_output.iterrows():
            instance_features = {}
            try:
                # --- a) Categorical Features ---
                # These names MUST align with your feature_names_for_prediction.json
                # Map from columns in df_output (e.g., STD_COL_PROCESS) to expected model input names.
                # Example: Subeje is a raw column, SubAxis is what the model expects
                instance_features['SubAxis'] = str(row.get('Subeje', 'Unknown_SubAxis'))
                instance_features['ProcessName'] = str(row.get(STD_COL_PROCESS, 'Unknown_Process'))
                instance_features['Periodicity'] = str(row.get(STD_COL_PERIODICITY, 'Unknown_Periodicity')).capitalize()
                instance_features['Status_D_minus_1'] = str(row.get(STD_COL_STATUS_YESTERDAY_RAW, 'UNKNOWN')).upper()


                # --- b) Numerical Features ---
                report_date_obj = row.get(STD_COL_REPORT_DATE)
                if isinstance(report_date_obj, str): report_date_obj = pd.to_datetime(report_date_obj, errors='coerce').date()
                elif isinstance(report_date_obj, datetime): report_date_obj = report_date_obj.date()
                # If still not a date object or is NaT, numerical date-dependent features will be default
                is_valid_date = isinstance(report_date_obj, dt_date) and not pd.isna(report_date_obj)

                commitment_time_str = row.get(STD_COL_TIME_EXPECTED_STR)
                commitment_time_obj = parse_time_robust(commitment_time_str) if pd.notna(commitment_time_str) else None

                instance_features['CommitmentHour'] = commitment_time_obj.hour if commitment_time_obj else -1
                instance_features['DayOfWeek'] = report_date_obj.weekday() if is_valid_date else -1

                # Calculate 'ObservedDurationMetric'
                current_status_today = str(row.get(STD_COL_STATUS_TODAY, 'UNKNOWN')).upper() # From df_output
                actual_time_str = row.get(STD_COL_TIME_ACTUAL_STR) # From df_output
                actual_time_obj = parse_time_robust(actual_time_str) if pd.notna(actual_time_str) else None
                obs_duration_metric = 0.0 # Default

                if is_valid_date:
                    if current_status_today == 'OK':
                        delay = calculate_delay(report_date_obj, commitment_time_obj, actual_time_obj)
                        obs_duration_metric = delay if pd.notna(delay) else 0.0
                    elif current_status_today == 'PENDING':
                        monitoring_time_obj = dt_time(MONITORING_HOUR, MONITORING_MINUTE)
                        if commitment_time_obj:
                            commitment_datetime = datetime.combine(report_date_obj, commitment_time_obj)
                            monitoring_datetime = datetime.combine(report_date_obj, monitoring_time_obj)
                            if monitoring_datetime > commitment_datetime:
                                obs_duration_metric = (monitoring_datetime - commitment_datetime).total_seconds() / 60.0
                            # else: obs_duration_metric remains 0.0
                        # else: obs_duration_metric remains 0.0
                    elif current_status_today == 'FAILED':
                        delay = calculate_delay(report_date_obj, commitment_time_obj, actual_time_obj)
                        obs_duration_metric = delay if pd.notna(delay) else 9999.0 # High value for failed if no time
                    # else: obs_duration_metric remains 0.0 (for UNKNOWN)
                else: # Invalid or missing report_date_obj
                    logging.debug(f"  AD Row {index}: Report_Date invalid, ObservedDurationMetric defaults for status {current_status_today}")
                    if current_status_today == 'FAILED': obs_duration_metric = 9999.0
                    # else remains 0.0

                instance_features['ObservedDurationMetric'] = obs_duration_metric

                # Ensure all features from EXPECTED_FEATURE_NAMES are present in instance_features
                # This loop creates the final dictionary with the exact expected keys.
                final_instance_for_prediction = {}
                for feat_name in EXPECTED_FEATURE_NAMES:
                    if feat_name in instance_features:
                        final_instance_for_prediction[feat_name] = instance_features[feat_name]
                    else:
                        # This means a feature your model expects was not engineered above.
                        # This indicates a mismatch between EXPECTED_FEATURE_NAMES and the engineering logic.
                        logging.error(f"  AD FATAL: Feature '{feat_name}' expected by model was not created for row {index}. Setting to NaN.")
                        final_instance_for_prediction[feat_name] = np.nan # Pipeline MUST handle this
                data_for_prediction_list.append(final_instance_for_prediction)

            except Exception as fe_error:
                logging.error(f"  AD: Error during feature engineering for row {index}: {fe_error}", exc_info=True)
                # Add a dummy entry to maintain length for prediction array alignment
                dummy_features = {feat: np.nan for feat in EXPECTED_FEATURE_NAMES}
                data_for_prediction_list.append(dummy_features)

        if not data_for_prediction_list:
            logging.warning(f"AD Skip for '{label}': No data to predict after feature engineering."); return df_output

        try:
            predictions_raw, probabilities_raw = predict_anomalies(data_for_prediction_list) # From ml_model.predictor
            if len(predictions_raw) == len(df_output):
                df_output['Is_Anomaly'] = [(pred == 0) for pred in predictions_raw] # Assuming 0=Anomaly, 1=Normal
                df_output['Anomaly_Probability_Normal'] = probabilities_raw
                num_anomalies = df_output['Is_Anomaly'].sum()
                logging.info(f"AD for '{label}' applied. Found {num_anomalies} anomalies out of {len(df_output)} instances.")
            else:
                logging.error(f"AD Mismatch: Predictions ({len(predictions_raw)}) vs DataFrame ({len(df_output)}). Anomalies not set.")
        except Exception as e:
            logging.error(f"Error during predict_anomalies call for '{label}': {e}", exc_info=True)
        return df_output


    def _apply_processing_rules(self):
        """Applies defined processing rules and then anomaly detection."""
        rules_are_present = bool(self.processing_rules) # Define it
        ad_is_possible = bool(self.scenario and self.scenario.enable_anomalies and ml_pipeline_from_predictor)

        if not rules_are_present and not ad_is_possible:
            logging.info("No processing rules and AD not possible/enabled. Skipping.")
            return
        logging.info("Applying processing rules and then anomaly detection (if enabled & model loaded)...")

        for label, data in list(self.results.items()): # Use list for safe iteration if modifying dict
            if isinstance(data, pd.DataFrame) and not data.empty:
                df_current = data.copy() # Work on a copy

                if rules_are_present:
                    logging.info(f"Applying rules for extracted DataFrame '{label}'...")
                    # ... (Your existing rule application logic that modifies df_current) ...
                    logging.info(f"Finished applying rules for '{label}'.")

                # Apply Anomaly Detection (will run if model loaded and scenario enabled)
                df_current = self._apply_anomaly_detection(df_current, label)

                self.results[label] = df_current # Update results dict with processed DataFrame
            elif isinstance(data, pd.DataFrame) and data.empty:
                 logging.info(f"DataFrame '{label}' is empty. Skipping rules and AD.")
                 # Ensure Is_Anomaly and Anomaly_Probability_Normal columns exist even for empty DFs if expected downstream
                 if 'Is_Anomaly' not in data.columns: data['Is_Anomaly'] = []
                 if 'Anomaly_Probability_Normal' not in data.columns: data['Anomaly_Probability_Normal'] = []
                 self.results[label] = data # Put back the empty (but potentially schema-consistent) DF

    def _save_monitoring_results(self):
        """
        Saves key features, ML predictions, and contextual information for each
        processed instance (row) to the MonitoringResult table.
        """
        if not self.log_entry:
            logging.error("Cannot save monitoring results: ExecutionLog entry not found."); return
        if not self.results:
            logging.info("No results extracted to save."); return

        logging.info(f"Saving key monitoring results for Log ID: {self.log_entry.log_id}...")
        results_to_add = []
        current_time = datetime.utcnow() # Use one timestamp for all results of this run

        # Find the primary DataFrame that went through AD (assume one main table)
        primary_df_label = None
        processed_df_for_saving = None
        for label, data in self.results.items():
            if isinstance(data, pd.DataFrame) and 'Is_Anomaly' in data.columns: # Identify by presence of AD output
                processed_df_for_saving = data
                primary_df_label = label
                logging.info(f"Found processed DataFrame '{primary_df_label}' for detailed result saving.")
                break
        
        if processed_df_for_saving is None:
            logging.warning("No processed DataFrame with anomaly flags found in results. Detailed metrics not saved.")
            # Optionally, still save scalar results if any
            for label, data in self.results.items():
                if not isinstance(data, pd.DataFrame): # Scalar results
                     results_to_add.append(MonitoringResult(
                         log_id=self.log_entry.log_id, metric_name=label, metric_value=str(data),
                         is_anomaly=False, recorded_at=current_time ))
            # ... (bulk save logic for scalar only) ...
            return

        # Iterate through each row of the processed DataFrame
        for index, row_data in processed_df_for_saving.iterrows():
            # Construct a base metric name prefix for this process instance
            # Uses the 'Process' column which should be standardized by now
            process_name_str = str(row_data.get(STD_COL_PROCESS, f"Row_{index}")).replace(" ", "_")
            periodicity_str = str(row_data.get(STD_COL_PERIODICITY, "UnknownP")).replace(" ", "_")
            base_metric_prefix = f"{primary_df_label}_{process_name_str}_{periodicity_str}_{index}" # Unique enough

            # 1. Save the overall anomaly flag for this instance/row
            is_row_anomalous = bool(row_data.get('Is_Anomaly', False))
            results_to_add.append(MonitoringResult(
                log_id=self.log_entry.log_id,
                metric_name=f"{base_metric_prefix}_IsAnomalousFlag", # Specific name for the flag
                metric_value=str(is_row_anomalous),
                is_anomaly=is_row_anomalous, # The flag itself
                anomaly_score=row_data.get('Anomaly_Probability_Normal'), # P(Normal)
                recorded_at=current_time
            ))

            # 2. Save key features that went into the model
            # These names MUST match the keys in your EXPECTED_FEATURE_NAMES / instance_features
            features_to_save = {
                'ObservedDurationMetric': row_data.get('ObservedDurationMetric'), # This is already in df_output
                'CommitmentHour': row_data.get('CommitmentHour_ModelInput'), # Assuming you named it this in feature engineering
                'DayOfWeek': row_data.get('DayOfWeek_ModelInput'),       # Assuming you named it this
                # Add any other key numerical features from EXPECTED_FEATURE_NAMES
            }
            # We also need Status_D_minus_1 for context if it was a feature
            if 'Status_D_minus_1' in EXPECTED_FEATURE_NAMES:
                 features_to_save['Status_D_minus_1'] = row_data.get('Status_D_minus_1') # From feature engineering

            for feat_name, feat_val in features_to_save.items():
                if pd.notna(feat_val): # Only save if value exists
                    results_to_add.append(MonitoringResult(
                        log_id=self.log_entry.log_id,
                        metric_name=f"{base_metric_prefix}_{feat_name}",
                        metric_value=str(feat_val),
                        is_anomaly=is_row_anomalous, # Associate with the row's anomaly status
                        anomaly_score=row_data.get('Anomaly_Probability_Normal'),
                        recorded_at=current_time
                    ))

            # 3. Save contextual "Typical_OK_Completion_Duration"
            # Construct key for lookup (ProcessName + Periodicity)
            process_identifier_for_lookup = f"{row_data.get(STD_COL_PROCESS, '')} ({row_data.get(STD_COL_PERIODICITY, '')})"
            typical_duration = self.typical_durations.get(process_identifier_for_lookup)
            if typical_duration is not None:
                results_to_add.append(MonitoringResult(
                    log_id=self.log_entry.log_id,
                    metric_name=f"{base_metric_prefix}_TypicalOKDuration",
                    metric_value=str(typical_duration),
                    is_anomaly=is_row_anomalous,
                    anomaly_score=row_data.get('Anomaly_Probability_Normal'),
                    recorded_at=current_time
                ))

            # 4. Save original Status_D for this row (for breakdown chart)
            original_status_d = str(row_data.get(STD_COL_STATUS_TODAY, 'UNKNOWN')) # STD_COL_STATUS_TODAY is the standardized 'OK'/'PENDING'/'FAILED'
            results_to_add.append(MonitoringResult(
                log_id=self.log_entry.log_id,
                metric_name=f"{base_metric_prefix}_OriginalStatusD",
                metric_value=original_status_d,
                is_anomaly=is_row_anomalous,
                anomaly_score=row_data.get('Anomaly_Probability_Normal'),
                recorded_at=current_time
            ))
            
            # 5. Save raw time strings for display on dashboard if needed
            results_to_add.append(MonitoringResult(log_id=self.log_entry.log_id, metric_name=f"{base_metric_prefix}_{STD_COL_TIME_EXPECTED_STR}", metric_value=str(row_data.get(STD_COL_TIME_EXPECTED_STR, '')), is_anomaly=is_row_anomalous, recorded_at=current_time))
            results_to_add.append(MonitoringResult(log_id=self.log_entry.log_id, metric_name=f"{base_metric_prefix}_{STD_COL_TIME_ACTUAL_STR}", metric_value=str(row_data.get(STD_COL_TIME_ACTUAL_STR, '')), is_anomaly=is_row_anomalous, recorded_at=current_time))


        # Save any scalar results from self.results as before
        for label, data in self.results.items():
            if not isinstance(data, pd.DataFrame): # Scalar results
                 results_to_add.append(MonitoringResult(
                     log_id=self.log_entry.log_id, metric_name=label, metric_value=str(data),
                     is_anomaly=False, recorded_at=current_time ))


        if results_to_add:
            try:
                db.session.bulk_save_objects(results_to_add)
                db.session.commit()
                logging.info(f"Successfully saved {len(results_to_add)} detailed monitoring results.")
            except Exception as e:
                 db.session.rollback()
                 logging.error(f"Error bulk saving detailed monitoring results: {e}", exc_info=True)
        else:
            logging.info("No detailed results formatted for saving.")


    def _generate_report(self):
        """
        Generates the Excel report using a template, scenario-level column mappings
        for tables, and step-level cell mappings for individual extracted elements.
        """
        if not self.log_entry:
            logging.error("Cannot generate report: ExecutionLog entry not found.")
            return

        if not self.scenario or not self.scenario.template or not self.scenario.template.file_path:
            logging.info("No report template configured or template path missing. Skipping report generation.")
            return

        template_path = self.scenario.template.file_path
        logging.info(f"Attempting to use template path from DB for report: {template_path}")

        if not os.path.exists(template_path):
            logging.error(f"Report template file not found at path: {template_path}")
            self.log_entry.log_message = (self.log_entry.log_message or "") + f"\nERROR: Template file not found at {template_path}"
            return

        logging.info(f"Generating report using template: {template_path}")
        try:
            workbook = openpyxl.load_workbook(template_path)
            # Assume data goes into the first/active sheet by default
            # This could be made configurable later if needed
            sheet = workbook.active
            logging.info(f"Using sheet: '{sheet.title}' for report generation.")

            # --- 1. Process EXTRACT_TABLE results using scenario-level ColumnMappings ---
            # Find the first DataFrame in results (assuming one main table for now)
            report_df_source_label = None
            report_df = None
            for label, data in self.results.items():
                if isinstance(data, pd.DataFrame) and not data.empty:
                    report_df = data
                    report_df_source_label = label
                    logging.info(f"Found DataFrame '{report_df_source_label}' with {len(report_df)} rows to use for main table report.")
                    break # Process first non-empty DataFrame encountered

            if report_df is not None:
                scenario_column_map_objects = self.scenario.column_mappings.all()
                if not scenario_column_map_objects:
                    logging.warning(f"No scenario-level column mappings defined for table data '{report_df_source_label}'. Table data might not be written as expected.")
                else:
                    column_map = {mapping.scraped_header: mapping.template_header for mapping in scenario_column_map_objects}
                    logging.debug(f"Applying column mappings for table: {column_map}")

                    # Prepare data based on scenario-level mappings
                    template_headers_ordered = [column_map[h] for h in report_df.columns if h in column_map] # Order by mapped headers found
                    mapped_data = pd.DataFrame()
                    for scraped_col, template_col_name in column_map.items():
                        if scraped_col in report_df.columns:
                            mapped_data[template_col_name] = report_df[scraped_col]
                        else:
                            logging.debug(f"  Table mapping: Scraped column '{scraped_col}' not in DataFrame '{report_df_source_label}'.")
                            # mapped_data[template_col_name] = None # Optionally add empty column

                    # Ensure columns are in the order they appear in the mappings (if desired)
                    # Or write based on template header existence.
                    # For simplicity, we iterate dataframe_to_rows on the potentially reordered mapped_data
                    if not mapped_data.empty:
                        # Try to reorder based on the order of values in column_map if needed
                        # For now, dataframe_to_rows will use mapped_data's current column order.
                        # If template_headers_ordered is important for specific placement:
                        # mapped_data_final = mapped_data[[h for h in template_headers_ordered if h in mapped_data.columns]]

                        # Find start row (e.g., after existing headers in template, or first empty)
                        # A common approach is to find the first empty row after row 1 (header)
                        start_row = 1
                        while sheet.cell(row=start_row, column=1).value is not None:
                            start_row += 1
                        if start_row == 1 and sheet.max_row >= 1: # If sheet has content but first data row is 1
                            start_row = sheet.max_row + 1

                        logging.info(f"Writing table data from '{report_df_source_label}' to sheet '{sheet.title}' starting from row {start_row}")

                        # Write rows using openpyxl utility (writes data only, not headers)
                        # Assuming template already has headers. If not, write mapped_data.columns first.
                        for r_idx, row_values in enumerate(dataframe_to_rows(mapped_data, index=False, header=False), start=start_row):
                            for c_idx, value in enumerate(row_values, start=1):
                                sheet.cell(row=r_idx, column=c_idx, value=value)
                        logging.info(f"  Wrote {len(mapped_data)} rows of table data.")
                    else:
                         logging.info(f"  No data from '{report_df_source_label}' matched scenario-level column mappings.")
            else:
                logging.info("No DataFrame results found to populate main table section of report.")


            # --- 2. Process EXTRACT_ELEMENT results with direct cell mapping ---
            # Fetch steps that have a specific cell mapping target
            steps_with_cell_mapping = ScenarioStep.query.filter(
                ScenarioStep.scenario_id == self.scenario.scenario_id,
                ScenarioStep.mapping_target_cell.isnot(None),
                ScenarioStep.mapping_target_cell != '', # Ensure it's not an empty string
                ScenarioStep.action_type == ActionTypeEnum.EXTRACT_ELEMENT
            ).all()

            if steps_with_cell_mapping:
                logging.info(f"Processing {len(steps_with_cell_mapping)} EXTRACT_ELEMENT steps with direct cell mappings...")
            for step_info in steps_with_cell_mapping:
                # The 'value' of the EXTRACT_ELEMENT step was used as the 'label' in self.results
                result_label = step_info.value # This is the key in self.results
                target_cell = step_info.mapping_target_cell.upper() # e.g., "A5"

                if result_label in self.results:
                    cell_value_to_write = self.results[result_label]
                    if isinstance(cell_value_to_write, pd.DataFrame):
                        logging.warning(f"  Step {step_info.sequence_order} (Label: '{result_label}') extracted a DataFrame but target '{target_cell}' is a single cell. Skipping this mapping.")
                        continue

                    # Validate cell reference format (basic check)
                    if re.match(r"^[A-Z]+[1-9]\d*$", target_cell):
                        try:
                            sheet[target_cell] = cell_value_to_write
                            logging.info(f"  Wrote value '{str(cell_value_to_write)[:50]}...' to cell {target_cell} (from step {step_info.sequence_order} mapping).")
                        except Exception as cell_write_error:
                            logging.error(f"  Error writing value to cell '{target_cell}' for step {step_info.sequence_order}: {cell_write_error}")
                    else:
                        logging.warning(f"  Invalid target cell format '{target_cell}' for step {step_info.sequence_order}. Skipping cell mapping.")
                else:
                    logging.warning(f"  No result found in self.results for label '{result_label}' (from step {step_info.sequence_order} mapping_target_cell).")


            # --- 3. Save the populated workbook ---
            output_dir = Path(self.app.config.get('REPORT_OUTPUT_DIR', './generated_reports')) # Use app config
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.scenario.custom_report_base_name:
                # Sanitize the custom base name slightly
                base_name = "".join(c if c.isalnum() or c in [' ', '_', '-'] else '_' for c in self.scenario.custom_report_base_name).strip()
                base_name = base_name.replace(' ', '_') # Replace spaces with underscores
                if not base_name: # If custom name becomes empty after sanitizing, use scenario name
                    base_name = "".join(c if c.isalnum() else "_" for c in self.scenario.name)
            else:
                # Default to scenario name if no custom base name
                base_name = "".join(c if c.isalnum() else "_" for c in self.scenario.name)
            # --- END CUSTOM BASE NAME ---

            output_filename = f"{base_name}_{timestamp}.xlsx" # Construct final filename
            output_path = output_dir / output_filename
            output_path_str = str(output_path.resolve())

            try:
                workbook.save(output_path) # 'workbook' is from previous openpyxl logic
                logging.info(f"Report saved successfully to: {output_path_str}")
                if self.log_entry:
                    self.log_entry.report_file_path = output_path_str
            except Exception as e:
                logging.error(f"Error saving report to {output_path_str}: {e}", exc_info=True)
                if self.log_entry:
                    self.log_entry.log_message = (self.log_entry.log_message or "") + f"\nERROR: Report saving failed: {e}"

        except Exception as e:
            logging.error(f"Error generating report: {e}", exc_info=True)
            if self.log_entry:
                self.log_entry.log_message = (self.log_entry.log_message or "") + f"\nERROR: Report generation failed: {e}"

    def _deliver_report(self):
        """Sends the generated report via email if configured."""
        # 1. Check if report exists and log entry is available
        if not self.log_entry or not self.log_entry.report_file_path:
            logging.info("No report file generated or log entry missing. Skipping email delivery.")
            return

        # 2. Check if email recipients are configured for the scenario
        if not self.scenario or not self.scenario.email_recipients:
            logging.info("No email recipients configured for this scenario. Skipping email delivery.")
            return

        # 3. Parse recipients string (comma-separated) into a list
        try:
            recipients = [email.strip() for email in self.scenario.email_recipients.split(',') if email.strip()]
            if not recipients:
                logging.warning("Recipient list parsed is empty. Skipping email delivery.")
                return
        except Exception as e:
            logging.error(f"Error parsing recipient list '{self.scenario.email_recipients}': {e}")
            return

        # 4. Verify the report file actually exists
        report_path = self.log_entry.report_file_path
        if not os.path.exists(report_path):
             logging.error(f"Cannot email report: File not found at {report_path}")
             self.log_entry.log_message = (self.log_entry.log_message or "") + f"\nERROR: Report file {report_path} not found for emailing."
             return

        logging.info(f"Attempting to email report '{os.path.basename(report_path)}' to: {recipients}")

        try:
            # 5. Construct the email message object
            subject = f"Monitoring Report: {self.scenario.name} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            # Get sender from the stored app instance's config
            sender = self.app.config.get('MAIL_DEFAULT_SENDER', self.app.config.get('MAIL_USERNAME'))
            if not sender:
                 logging.error("MAIL_DEFAULT_SENDER or MAIL_USERNAME not configured in Flask app. Cannot send email.")
                 self.log_entry.log_message = (self.log_entry.log_message or "") + "\nERROR: Mail sender not configured."
                 return

            msg = Message(subject=subject,
                          sender=sender,
                          recipients=recipients)

            # Construct email body
            msg.body = f"Automated monitoring report for scenario '{self.scenario.name}' attached.\n\n"
            msg.body += f"Execution completed at: {self.log_entry.end_time.strftime('%Y-%m-%d %H:%M:%S UTC') if self.log_entry.end_time else 'N/A'}\n"
            msg.body += f"Status: {self.log_entry.status.name}\n"
            # Include any previous log messages (like errors or warnings)
            if self.log_entry.log_message and "Report emailed successfully" not in self.log_entry.log_message and "ERROR: Failed to send report email" not in self.log_entry.log_message :
                 msg.body += f"\nNotes/Errors during execution:\n{self.log_entry.log_message}\n"

            # 6. Attach the report file
            with open(report_path, 'rb') as fp:
                msg.attach(
                    filename=os.path.basename(report_path),
                    # Correct MIME type for .xlsx files
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    data=fp.read()
                )

            # 7. Send the email using the mail instance from the stored app instance
            mail_instance = self.app.extensions.get('mail')
            if mail_instance:
                 # Ensure app context is active for sending email
                 with self.app.app_context():
                      mail_instance.send(msg)
                 logging.info(f"Report email sent successfully to {recipients}.")
                 # Update log message - Append success
                 self.log_entry.log_message = (self.log_entry.log_message or "") + "\nINFO: Report emailed successfully."
            else:
                 # This indicates a setup problem if Flask-Mail wasn't initialized correctly
                 logging.error("Flask-Mail extension not found in provided app instance. Mail not sent.")
                 self.log_entry.log_message = (self.log_entry.log_message or "") + "\nERROR: Flask-Mail extension not found in app."
                 # Maybe raise error if this happens?

        except Exception as e:
            # Catch any exception during mail sending (e.g., auth error, connection error)
            logging.error(f"Failed to send report email: {e}", exc_info=True)
            # Update log message with the specific email error
            self.log_entry.log_message = (self.log_entry.log_message or "") + f"\nERROR: Failed to send report email: {e}"
            # Note: We don't re-raise the exception here typically,
            # as failure to email shouldn't necessarily fail the whole monitoring run.


    


    def run(self):
        """Executes the full scenario."""
        final_status = ExecutionStatusEnum.FAILED
        error_message = None
        try:
            self._setup()
            if not self.page: raise RuntimeError("Playwright page not initialized correctly.")
            # self._perform_login() # Call login method if needed for the target app

            steps = self.scenario.steps.order_by(ScenarioStep.sequence_order).all()
            if not steps:
                logging.warning(f"Scenario {self.scenario_id} has no steps defined.")
                final_status, error_message = ExecutionStatusEnum.SUCCESS, "Scenario completed (No steps)."
            else:
                 for step in steps: self._execute_step(step)
                 self._apply_processing_rules() # Applies rules & anomaly detection structure
                 self._save_monitoring_results() # Saves results including anomaly flags
                 self._generate_report() # Generates report possibly including anomaly flags

                 self._deliver_report()

                 logging.info("Final processed results dictionary (after AD):")
                 for label, result_data in self.results.items():
                      if isinstance(result_data, pd.DataFrame): logging.info(f"  Result '{label}' (DataFrame): {len(result_data)} rows x {len(result_data.columns)} columns")
                      else: logging.info(f"  Result '{label}': {str(result_data)[:200]}...")

                 

                 logging.info(f"Scenario ID {self.scenario_id} execution finished.")
                 final_status, error_message = ExecutionStatusEnum.SUCCESS, "Scenario executed successfully."

        except PlaywrightTimeoutError as e: error_message = f"Timeout Error: {e}"; logging.error(error_message, exc_info=False)
        except Exception as e: error_message = f"Error: {e}"; logging.error(error_message, exc_info=True)
        finally:
            if final_status == ExecutionStatusEnum.FAILED and self.page and self.browser and self.browser.is_connected():
                 try:
                    screenshot_path = f"error_screenshot_scenario_{self.scenario_id}_log_{self.log_entry.log_id if self.log_entry else 'unknown'}.png"
                    self.page.screenshot(path=screenshot_path, full_page=True)
                    logging.info(f"Error screenshot saved to {screenshot_path}")
                    error_message = (error_message or "") + f" (Screenshot: {screenshot_path})"
                 except Exception as se: logging.error(f"Failed to take error screenshot: {se}")
            self._teardown(final_status, error_message)
        return final_status == ExecutionStatusEnum.SUCCESS